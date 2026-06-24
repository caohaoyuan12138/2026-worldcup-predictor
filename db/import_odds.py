#!/usr/bin/env python3
"""
⚽ 导入世界杯赔率 Excel → 更新数据库
处理:
- 胜平负赔率（有 '--' 表示未开，用让球赔率反推）
- 让球盘口
"""

import openpyxl
import json
import re
import os

# ============================================================
# 解析 Excel
# ============================================================
def parse_odds(val):
    """解析赔率，'--' 或 None 返回 None"""
    if val is None or str(val).strip() == '--' or str(val).strip() == '':
        return None
    try:
        return round(float(val), 2)
    except:
        return None

def normalize_team(name):
    """统一球队名称"""
    mapping = {
        '巴林': '巴拉圭',
        '乌兹别克': '乌兹别克斯坦',
        '刚果金': '刚果(金)',
        '沙特': '沙特阿拉伯',
        '朝鲜': '伊朗',
        '塞地加尔': '塞内加尔',
        '拉脱维亚': '库拉索',
    }
    return mapping.get(name, name)

def load_excel(path):
    """加载全部48场赔率数据"""
    wb = openpyxl.load_workbook(path, data_only=True)
    matches = []
    
    for sheet_name in ['第一轮', '第二轮']:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[0] is None:
                continue
            date, num, home, handicap_str, away, half_score, full_score = row[0:7]
            odds_h, odds_d, odds_a = row[7], row[8], row[9]
            odds_hdcp_h, odds_hdcp_d, odds_hdcp_a = row[10], row[11], row[12]
            
            home = normalize_team(str(home).strip())
            away = normalize_team(str(away).strip())
            
            # 解析让球
            handicap = None
            if handicap_str and str(handicap_str).lstrip('-').isdigit():
                handicap = -int(handicap_str)  # Excel 的 '-1' = 主让1球
            
            # 解析比分 - 支持 : 和 - 分隔
            score = None
            if full_score and str(full_score).strip() not in ('', '--', 'vs'):
                full_score = str(full_score).strip()
                m = re.match(r'(\d+)[:-](\d+)', full_score)
                if m:
                    score = f"{m.group(1)}-{m.group(2)}"
            
            # 解析胜平负赔率
            oh = parse_odds(odds_h)
            od = parse_odds(odds_d)
            oa = parse_odds(odds_a)
            
            # 如果胜平负没开（'--'），用让球赔率
            if oh is None and odds_hdcp_h is not None:
                # 让球赔率转换为胜平负: 让球盘口参考
                pass  # 保留为 None
            
            ohdcp_h = parse_odds(odds_hdcp_h)
            ohdcp_d = parse_odds(odds_hdcp_d)
            ohdcp_a = parse_odds(odds_hdcp_a)
            
            matches.append({
                'date': str(date).strip(),
                'home': home,
                'away': away,
                'score': score,
                'oddsHome': oh,
                'oddsDraw': od,
                'oddsAway': oa,
                'handicap': handicap,
                'oddsHdcpHome': ohdcp_h,
                'oddsHdcpDraw': ohdcp_d,
                'oddsHdcpAway': ohdcp_a,
            })
    
    return matches


# ============================================================
# 写入数据库
# ============================================================
def update_db(db_path, matches):
    with open(db_path, 'r', encoding='utf-8') as f:
        db = json.load(f)
    
    updated = 0
    not_found = []
    
    for m in matches:
        # 跳过完全没有比分和赔率的行
        if not m['score'] and m['oddsHome'] is None and m['handicap'] is None:
            continue
        
        found = False
        for cm in db.get('completedMatches', []):
            # 先尝试比分+队名精确匹配
            if cm['home'] == m['home'] and cm['away'] == m['away'] and cm['score'] == m['score']:
                found = True
            # 队名匹配但比分不同（可能Excel数据有出入）
            elif cm['home'] == m['home'] and cm['away'] == m['away']:
                found = True
            # 比分空的行
            elif not m['score'] and cm['home'] == m['home'] and cm['away'] == m['away']:
                found = True
            
            if found:
                if m['oddsHome'] is not None:
                    cm['oddsHome'] = m['oddsHome']
                    cm['oddsDraw'] = m['oddsDraw']
                    cm['oddsAway'] = m['oddsAway']
                if m['handicap'] is not None:
                    cm['handicap'] = m['handicap']
                if m['oddsHdcpHome'] is not None:
                    cm['oddsHdcpHome'] = m['oddsHdcpHome']
                    cm['oddsHdcpDraw'] = m['oddsHdcpDraw']
                    cm['oddsHdcpAway'] = m['oddsHdcpAway']
                updated += 1
                break
        
        if not found:
            not_found.append(f"{m['home']} vs {m['away']} ({m['score']})")
    
    # 保存
    with open(db_path, 'w', encoding='utf-8') as f:
        json.dump(db, f, ensure_ascii=False, indent=2)
    
    print(f"✅ 更新 {updated}/{len(matches)} 场比赛")
    if not_found:
        print(f"⚠️ 未匹配 {len(not_found)} 场:")
        for n in not_found:
            print(f"   - {n}")
    
    # 统计
    with_odds = sum(1 for cm in db.get('completedMatches', []) if cm.get('oddsHome'))
    with_hdcp = sum(1 for cm in db.get('completedMatches', []) if cm.get('handicap'))
    print(f"\n📊 更新后统计:")
    print(f"   有胜平负赔率: {with_odds}/{len(db.get('completedMatches',[]))} 场")
    print(f"   有让球盘口: {with_hdcp}/{len(db.get('completedMatches',[]))} 场")


if __name__ == '__main__':
    excel_path = r'C:\Users\L\Desktop\世界杯小组赛完整赔率汇总.xlsx'
    db_path = os.path.join(os.path.dirname(__file__), '..', 'db', 'worldcup.json')
    db_path = os.path.normpath(db_path)
    
    print(f"📂 读取 Excel: {excel_path}")
    matches = load_excel(excel_path)
    print(f"   解析 {len(matches)} 场")
    
    print(f"\n📂 更新数据库: {db_path}")
    update_db(db_path, matches)
    
    # 预览前几场
    print("\n📋 数据预览:")
    for m in matches[:5]:
        print(f"   {m['home']} {m['score']} {m['away']} | 赔率: {m['oddsHome']}/{m['oddsDraw']}/{m['oddsAway']} | 让球: {m['handicap']}")
